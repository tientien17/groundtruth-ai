"""Tests for quantity Excel export."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from uuid import UUID, uuid4

from fastapi.testclient import TestClient
from openpyxl import load_workbook  # type: ignore[reportMissingModuleSource]

from database import get_session, init_db
from main import create_app
from models import Classification, Document, Project, Sheet, TakeoffItem


PROJECT_ID = UUID("00000000-0000-0000-0000-000000000016")
DOCUMENT_ID = UUID("00000000-0000-0000-0000-000000000116")
SHEET_ONE_ID = UUID("00000000-0000-0000-0000-000000001016")
SHEET_TWO_ID = UUID("00000000-0000-0000-0000-000000002016")
WALL_CLASS_ID = UUID("00000000-0000-0000-0000-000000010016")
DOOR_CLASS_ID = UUID("00000000-0000-0000-0000-000000020016")


def create_export_project(tmp_path: Path) -> Path:
    project_path = tmp_path / "Export.gtl"
    project_path.mkdir()
    db_path = project_path / "project.sqlite"
    init_db(db_path)

    with get_session(db_path) as session:
        session.add(Project(id=PROJECT_ID, name="Export", path=str(project_path)))
        session.add(
            Document(
                id=DOCUMENT_ID,
                project_id=PROJECT_ID,
                filename="plans.pdf",
                original_path=str(project_path / "plans.pdf"),
            )
        )
        session.add_all(
            [
                Sheet(
                    id=SHEET_ONE_ID,
                    document_id=DOCUMENT_ID,
                    sheet_number="A-101",
                    sheet_title="Floor Plan",
                    page_index=0,
                ),
                Sheet(
                    id=SHEET_TWO_ID,
                    document_id=DOCUMENT_ID,
                    sheet_number="A-102",
                    sheet_title="Ceiling Plan",
                    page_index=1,
                ),
            ]
        )
        session.add_all(
            [
                Classification(
                    id=WALL_CLASS_ID,
                    project_id=PROJECT_ID,
                    name="Walls",
                    color="#2563EB",
                    unit="ft",
                ),
                Classification(
                    id=DOOR_CLASS_ID,
                    project_id=PROJECT_ID,
                    name="Doors",
                    color="#16A34A",
                    unit="count",
                ),
            ]
        )
        session.add_all(
            [
                TakeoffItem(
                    id=uuid4(),
                    sheet_id=SHEET_ONE_ID,
                    classification_id=WALL_CLASS_ID,
                    type="linear",
                    geometry_json={},
                    source="manual",
                    quantity_raw=100,
                    quantity_unit="ft",
                    formulas={"waste": "QTY * 1.1"},
                    created_by="tester",
                ),
                TakeoffItem(
                    id=uuid4(),
                    sheet_id=SHEET_ONE_ID,
                    classification_id=DOOR_CLASS_ID,
                    type="count",
                    geometry_json={},
                    source="manual",
                    quantity_raw=4,
                    quantity_unit="count",
                    formulas={},
                    created_by="tester",
                ),
                TakeoffItem(
                    id=uuid4(),
                    sheet_id=SHEET_TWO_ID,
                    classification_id=WALL_CLASS_ID,
                    type="linear",
                    geometry_json={},
                    source="manual",
                    quantity_raw=55,
                    quantity_unit="ft",
                    formulas={},
                    created_by="tester",
                ),
            ]
        )

    return project_path


def test_export_project_xlsx_groups_quantities_across_sheets(tmp_path: Path):
    project_path = create_export_project(tmp_path)
    client = TestClient(create_app())

    response = client.get(
        f"/projects/{PROJECT_ID}/export.xlsx",
        params={"project_path": str(project_path)},
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["A-101 Floor Plan", "A-102 Ceiling Plan"]

    first_sheet = workbook["A-101 Floor Plan"]
    assert [cell.value for cell in first_sheet[1]] == [
        "Classification",
        "Raw Qty",
        "Unit",
        "Formula",
        "Final Qty",
    ]
    assert [first_sheet.cell(2, column).value for column in range(1, 6)] == [
        "Doors",
        4,
        "count",
        None,
        4,
    ]
    assert [first_sheet.cell(3, column).value for column in range(1, 6)] == [
        "Walls",
        100,
        "ft",
        "QTY * 1.1",
        110,
    ]


def test_export_sheet_xlsx_limits_workbook_to_requested_sheet(tmp_path: Path):
    project_path = create_export_project(tmp_path)
    client = TestClient(create_app())

    response = client.get(
        f"/projects/{PROJECT_ID}/export.xlsx",
        params={"project_path": str(project_path), "sheet_id": str(SHEET_TWO_ID)},
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["A-102 Ceiling Plan"]
    sheet = workbook["A-102 Ceiling Plan"]
    assert [sheet.cell(2, column).value for column in range(1, 6)] == [
        "Walls",
        55,
        "ft",
        None,
        55,
    ]
