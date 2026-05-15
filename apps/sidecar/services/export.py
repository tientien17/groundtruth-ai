"""Excel export helpers for takeoff quantities."""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import Workbook  # type: ignore[reportMissingModuleSource]
from openpyxl.styles import Font, PatternFill  # type: ignore[reportMissingModuleSource]
from sqlalchemy.orm import Session

from models import Classification, Document, Sheet, TakeoffItem
from services.formulas import FormulaError, apply_formula


EXPORT_COLUMNS = ["Classification", "Raw Qty", "Unit", "Formula", "Final Qty"]


@dataclass(frozen=True)
class QuantityExportRow:
    """Flattened row written to the Excel workbook."""

    classification: str
    raw_qty: float
    unit: str
    formula: str
    final_qty: float


def build_takeoff_export_workbook(
    session: Session,
    project_id: Any,
    sheet_id: Any | None = None,
) -> BytesIO:
    """Create an Excel workbook for all project sheets or one sheet."""
    sheets = _query_sheets(session, project_id, sheet_id)
    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is not None:
        workbook.remove(active_sheet)

    if not sheets:
        sheet = workbook.create_sheet("Quantities")
        _write_sheet(sheet, [])
    else:
        used_titles: set[str] = set()
        for project_sheet in sheets:
            title = _unique_sheet_title(_sheet_title(project_sheet), used_titles)
            worksheet = workbook.create_sheet(title)
            rows = build_quantity_rows(session, project_sheet.id)
            _write_sheet(worksheet, rows)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_quantity_rows(session: Session, sheet_id: Any) -> list[QuantityExportRow]:
    """Return takeoff rows grouped and ordered by classification for one sheet."""
    classifications = {
        classification.id: classification
        for classification in session.query(Classification).all()
    }
    items = (
        session.query(TakeoffItem)
        .filter(TakeoffItem.sheet_id == sheet_id)
        .order_by(TakeoffItem.created_at, TakeoffItem.id)
        .all()
    )

    grouped: dict[str, list[QuantityExportRow]] = defaultdict(list)
    for item in items:
        classification = classifications.get(item.classification_id) if item.classification_id else None
        name = classification.name if classification else "Unclassified"
        raw_qty = float(item.quantity_raw or 0)
        formula = _formula_expression(item.formulas or {})
        grouped[name].append(
            QuantityExportRow(
                classification=name,
                raw_qty=raw_qty,
                unit=item.quantity_unit or (classification.unit if classification else ""),
                formula=formula,
                final_qty=_final_quantity(raw_qty, formula),
            )
        )

    rows: list[QuantityExportRow] = []
    for classification_name in sorted(grouped, key=str.casefold):
        rows.extend(grouped[classification_name])
    return rows


def _query_sheets(session: Session, project_id: Any, sheet_id: Any | None) -> list[Sheet]:
    query = (
        session.query(Sheet)
        .join(Sheet.document)
        .filter(Document.project_id == project_id)
        .order_by(Sheet.page_index, Sheet.sheet_number)
    )
    if sheet_id is not None:
        query = query.filter(Sheet.id == sheet_id)
    return query.all()


def _write_sheet(worksheet, rows: list[QuantityExportRow]) -> None:
    worksheet.append(EXPORT_COLUMNS)
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    current_classification: str | None = None
    for row in rows:
        if row.classification != current_classification:
            current_classification = row.classification
        worksheet.append([
            row.classification,
            row.raw_qty,
            row.unit,
            row.formula,
            row.final_qty,
        ])

    widths = [24, 12, 12, 24, 12]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(1, column_index).column_letter].width = width


def _formula_expression(formulas: dict[str, Any]) -> str:
    if not formulas:
        return ""
    if isinstance(formulas.get("expression"), str):
        return formulas["expression"]
    for value in formulas.values():
        if isinstance(value, str):
            return value
    return ""


def _final_quantity(raw_qty: float, formula: str) -> float:
    if not formula:
        return raw_qty
    try:
        return apply_formula(formula, raw_qty).value
    except FormulaError:
        return raw_qty


def _sheet_title(sheet: Sheet) -> str:
    if sheet.sheet_title:
        return f"{sheet.sheet_number} {sheet.sheet_title}"
    return sheet.sheet_number or "Quantities"


def _unique_sheet_title(title: str, used_titles: set[str]) -> str:
    safe = "".join("-" if char in r"[]:*?/\\" else char for char in title)[:31] or "Quantities"
    candidate = safe
    suffix = 2
    while candidate in used_titles:
        ending = f" {suffix}"
        candidate = f"{safe[:31 - len(ending)]}{ending}"
        suffix += 1
    used_titles.add(candidate)
    return candidate
